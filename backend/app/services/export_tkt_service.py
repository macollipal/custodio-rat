"""
Servicio de exportación de tickets ARCO a CSV, Excel y PDF.
ARCO-QW1: Exportación CSV/Excel/PDF de tickets.
"""
import csv
import io
from datetime import datetime, timezone
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy.orm import Session

from app.models.tkt_solicitud_derecho import TktSolicitudDerecho
from app.models.company import Company
from app.models.user import User


TIPO_LABELS = {
    "acceso": "Acceso (Art. 12)",
    "rectificacion": "Rectificación (Art. 12)",
    "cancelacion": "Cancelación (Art. 12)",
    "oposicion": "Oposición (Art. 12)",
    "bloqueo": "Bloqueo (Art. 12)",
    "portabilidad": "Portabilidad (Art. 9)",
}

ESTADO_LABELS = {
    "abierto": "Abierto",
    "en_proceso": "En Proceso",
    "pendiente": "Pendiente",
    "resuelto": "Resuelto",
    "bloqueado": "Bloqueado",
    "rechazado": "Rechazado",
    "subsanacion": "Subsanación",
    "prorroga": "Prórroga",
}

PRIORIDAD_LABELS = {
    "baja": "Baja",
    "normal": "Normal",
    "urgente": "Urgente",
}


def _get_tickets_para_exportar(
    db: Session,
    company_id: Optional[int],
    estado: Optional[str],
    prioridad: Optional[str],
    fecha_desde: Optional[str],
    fecha_hasta: Optional[str],
) -> list[TktSolicitudDerecho]:
    q = db.query(TktSolicitudDerecho)
    if company_id:
        q = q.filter(TktSolicitudDerecho.company_id == company_id)
    if estado:
        q = q.filter(TktSolicitudDerecho.estado == estado)
    if prioridad:
        q = q.filter(TktSolicitudDerecho.prioridad == prioridad)
    if fecha_desde:
        from datetime import datetime as dt
        try:
            dt_desde = dt.fromisoformat(fecha_desde).replace(tzinfo=timezone.utc)
            q = q.filter(TktSolicitudDerecho.fecha_recepcion >= dt_desde)
        except ValueError:
            pass
    if fecha_hasta:
        from datetime import datetime as dt
        try:
            dt_hasta = dt.fromisoformat(fecha_hasta).replace(tzinfo=timezone.utc)
            q = q.filter(TktSolicitudDerecho.fecha_recepcion <= dt_hasta)
        except ValueError:
            pass
    return q.order_by(TktSolicitudDerecho.fecha_recepcion.desc()).all()


def _enriquecer_ticket(ticket: TktSolicitudDerecho, db: Session) -> dict:
    empresa = db.query(Company).filter(Company.id == ticket.company_id).first()
    responsable = None
    if ticket.responsable_id:
        user = db.query(User).filter(User.id == ticket.responsable_id).first()
        if user:
            responsable = user.full_name or user.username

    from app.services.ticket_service import calcular_dias_restantes
    dias_rest = calcular_dias_restantes(ticket.fecha_vencimiento) if ticket.fecha_vencimiento else None

    return {
        "id": ticket.id,
        "empresa": empresa.nombre if empresa else str(ticket.company_id),
        "tipo": TIPO_LABELS.get(ticket.tipo, ticket.tipo),
        "estado": ESTADO_LABELS.get(ticket.estado, ticket.estado),
        "prioridad": PRIORIDAD_LABELS.get(ticket.prioridad, ticket.prioridad),
        "origen": ticket.origen,
        "titular_nombre": ticket.titular_nombre,
        "titular_email": ticket.titular_email,
        "titular_rut": ticket.titular_rut or "",
        "descripcion": (ticket.descripcion or "")[:200],
        "fecha_recepcion": ticket.fecha_recepcion.strftime("%d-%m-%Y %H:%M") if ticket.fecha_recepcion else "",
        "fecha_vencimiento": ticket.fecha_vencimiento.strftime("%d-%m-%Y") if ticket.fecha_vencimiento else "",
        "dias_restantes": dias_rest if dias_rest is not None else "",
        "responsable": responsable or "Sin asignar",
        "respuesta_texto": (ticket.respuesta_texto or "")[:200],
        "respuesta_fecha": ticket.respuesta_fecha.strftime("%d-%m-%Y") if ticket.respuesta_fecha else "",
        "representante_nombre": ticket.representante_nombre or "",
        "representante_rut": ticket.representante_rut or "",
        "telefono": ticket.telefono or "",
        "pais": ticket.pais or "",
        "created_at": ticket.created_at.strftime("%d-%m-%Y %H:%M") if ticket.created_at else "",
    }


HEADERS_CSV = [
    "ID", "Empresa", "Tipo", "Estado", "Prioridad", "Origen",
    "Titular Nombre", "Titular Email", "Titular RUT",
    "Descripción", "Fecha Recepción", "Fecha Vencimiento",
    "Días Restantes", "Responsable", "Respuesta", "Fecha Respuesta",
    "Representante", "RUT Rep.", "Teléfono", "País",
]


def generar_csv(
    db: Session,
    company_id: Optional[int],
    estado: Optional[str],
    prioridad: Optional[str],
    fecha_desde: Optional[str],
    fecha_hasta: Optional[str],
) -> bytes:
    tickets = _get_tickets_para_exportar(db, company_id, estado, prioridad, fecha_desde, fecha_hasta)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)
    writer.writerow(HEADERS_CSV)
    for t in tickets:
        row = _enriquecer_ticket(t, db)
        writer.writerow([
            row["id"], row["empresa"], row["tipo"], row["estado"], row["prioridad"], row["origen"],
            row["titular_nombre"], row["titular_email"], row["titular_rut"],
            row["descripcion"], row["fecha_recepcion"], row["fecha_vencimiento"],
            row["dias_restantes"], row["responsable"],
            row["respuesta_texto"], row["respuesta_fecha"],
            row["representante_nombre"], row["representante_rut"],
            row["telefono"], row["pais"],
        ])
    return output.getvalue().encode("utf-8-sig")


def generar_excel(
    db: Session,
    company_id: Optional[int],
    estado: Optional[str],
    prioridad: Optional[str],
    fecha_desde: Optional[str],
    fecha_hasta: Optional[str],
) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tickets = _get_tickets_para_exportar(db, company_id, estado, prioridad, fecha_desde, fecha_hasta)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets ARCO"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers_display = [
        "ID", "Empresa", "Tipo", "Estado", "Prioridad", "Origen",
        "Titular", "Email Titular", "RUT Titular",
        "Descripción", "Fecha Recepción", "Fecha Vencimiento",
        "Días Rest.", "Responsable", "Respuesta", "Fecha Respuesta",
        "Representante", "RUT Rep.", "Teléfono", "País",
    ]

    for col, h in enumerate(headers_display, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, t in enumerate(tickets, 2):
        row = _enriquecer_ticket(t, db)
        values = [
            row["id"], row["empresa"], row["tipo"], row["estado"], row["prioridad"], row["origen"],
            row["titular_nombre"], row["titular_email"], row["titular_rut"],
            row["descripcion"], row["fecha_recepcion"], row["fecha_vencimiento"],
            row["dias_restantes"], row["responsable"],
            row["respuesta_texto"], row["respuesta_fecha"],
            row["representante_nombre"], row["representante_rut"],
            row["telefono"], row["pais"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if row["dias_restantes"] != "":
                if row["dias_restantes"] <= 0:
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                elif row["dias_restantes"] <= 2:
                    cell.fill = PatternFill("solid", fgColor="FEF9C8")

    for col in range(1, len(headers_display) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers_display))}1"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generar_pdf(
    db: Session,
    company_id: Optional[int],
    estado: Optional[str],
    prioridad: Optional[str],
    fecha_desde: Optional[str],
    fecha_hasta: Optional[str],
) -> bytes:
    tickets = _get_tickets_para_exportar(db, company_id, estado, prioridad, fecha_desde, fecha_hasta)
    output = io.BytesIO()

    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CustomTitle", parent=styles["Heading1"], fontSize=14, spaceAfter=12, textColor=colors.HexColor("#1E40AF"))
    normal_style = styles["Normal"]

    elements = []
    elements.append(Paragraph("Reporte de Solicitudes ARCO — Custodio RAT Manager", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')} · Total: {len(tickets)} solicitudes", styles["Normal"]))
    elements.append(Spacer(1, 8*mm))

    if not tickets:
        elements.append(Paragraph("No se encontraron solicitudes con los filtros seleccionados.", styles["Normal"]))
        doc.build(elements)
        return output.getvalue()

    table_data = [
        ["ID", "Empresa", "Tipo", "Estado", "Prioridad", "Titular", "Vencimiento", "Días"],
    ]
    for t in tickets:
        from app.services.ticket_service import calcular_dias_restantes
        dias = calcular_dias_restantes(t.fecha_vencimiento) if t.fecha_vencimiento else 0
        empresa = db.query(Company).filter(Company.id == t.company_id).first()
        table_data.append([
            str(t.id),
            (empresa.nombre if empresa else str(t.company_id))[:20],
            TIPO_LABELS.get(t.tipo, t.tipo)[:15],
            ESTADO_LABELS.get(t.estado, t.estado)[:12],
            PRIORIDAD_LABELS.get(t.prioridad, t.prioridad),
            t.titular_nombre[:20],
            t.fecha_vencimiento.strftime("%d-%m-%Y") if t.fecha_vencimiento else "-",
            f"{dias}d" if dias is not None else "-",
        ])

    col_widths = [20*mm, 35*mm, 30*mm, 28*mm, 22*mm, 38*mm, 28*mm, 15*mm]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    for i, row in enumerate(table_data[1:], 1):
        dias_str = row[-1]
        if dias_str and dias_str != "-":
            dias_val = int(dias_str.replace("d", ""))
            if dias_val <= 0:
                table.setStyle(TableStyle([("BACKGROUND", (7, i), (7, i), colors.HexColor("#FEE2E2"))]))
            elif dias_val <= 2:
                table.setStyle(TableStyle([("BACKGROUND", (7, i), (7, i), colors.HexColor("#FEF9C8"))]))

    elements.append(table)
    doc.build(elements)
    return output.getvalue()
